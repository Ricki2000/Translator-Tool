import tkinter as tk
from tkinter import filedialog
from googletrans import Translator
import openpyxl

translator = Translator()

def translate_excel():
    file_path = filedialog.askopenfilename()
    if file_path:
        target_language = 'en'
        workbook = openpyxl.load_workbook(file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        translated_text = translator.translate(cell.value, dest=target_language)
                        cell.value = translated_text.text

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if save_path:
            workbook.save(save_path)

# Create the main window
root = tk.Tk()
root.title("Excel Translator")

# Create a button for selecting the Excel file
select_file_button = tk.Button(root, text="Select Excel File", command=translate_excel)
select_file_button.pack(pady=20)

# Run the main event loop
root.mainloop()