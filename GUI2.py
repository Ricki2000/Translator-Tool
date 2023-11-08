import tkinter as tk
from tkinter import ttk, filedialog
from googletrans import Translator
import openpyxl
import time

root = tk.Tk()
root.title("Excel Translator")

translator = Translator()
start_time = None

def translate_excel():
    global start_time
    start_time = time.time()

    file_path = file_path_var.get()

    if file_path:
        target_language = 'en'
        workbook = openpyxl.load_workbook(file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        translated_text = translator.translate(cell.value, dest=target_language)
                        cell.value = translated_text.text

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if save_path:
            workbook.save(save_path)

        end_time = time.time()
        elapsed_time = end_time - start_time
        clock_label.config(text=f"Time Elapsed: {elapsed_time:.2f} seconds")

def browse_file():
    filename = filedialog.askopenfilename(initialdir="/", title="Select file", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    file_path_var.set(filename)

file_path_var = tk.StringVar()

main_frame = ttk.Frame(root, padding=20)
main_frame.grid(row=0, column=0, sticky="nsew")

clock_label = ttk.Label(main_frame, text="Time Elapsed: 0 seconds")
clock_label.grid(row=0, column=0, columnspan=2, pady=10)

file_label = ttk.Label(main_frame, text="Selected Excel file:")
file_label.grid(row=1, column=0, pady=10)

file_entry = ttk.Entry(main_frame, textvariable=file_path_var, width=50)
file_entry.grid(row=1, column=1, pady=10)

browse_button = ttk.Button(main_frame, text="Browse", command=browse_file)
browse_button.grid(row=2, column=0, columnspan=2, pady=10)

translate_button = ttk.Button(main_frame, text="Translate Excel File", command=translate_excel)
translate_button.grid(row=3, column=0, columnspan=2, pady=10)

root.mainloop()
