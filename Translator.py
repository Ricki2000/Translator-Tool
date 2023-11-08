from googletrans import Translator 
import pandas as pd
import openpyxl

translator = Translator()

out = translator.translate("Na, wie geht es dir?", dest="en")

# Pfad zur Excel-Datei
excel_file_path = r"C:\Users\Ricardo.Jordan\OneDrive - Vodafone Group\Documents\Python\IMADtasks\Evaluation Insights & Features for App Product.xlsx"

# Zielsprache für die Übersetzung
target_language = 'en'

# Initialisiere den Translator
translator = Translator()

# Öffne die Excel-Datei
workbook = openpyxl.load_workbook(excel_file_path)

# Durchlaufe die Blätter der Excel-Datei
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    
    # Durchlaufe die Zellen in jedem Blatt
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                translated_text = translator.translate(cell.value, dest=target_language)
                cell.value = translated_text.text

# Speichere die geänderte Excel-Datei
workbook.save('übersetzte_datei.xlsx')