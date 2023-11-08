import tkinter as tk  # Import the tkinter module for creating GUI applications
from tkinter import ttk, filedialog  # Import specific components from tkinter for improved styling and file dialog functionality
from googletrans import Translator  # Import Translator from the googletrans module for language translation
import openpyxl  # Import the openpyxl module for working with Excel files
import datetime  # Import the datetime module for time-related operations
from fastapi import FastAPI
from pydantic import BaseModel


#pydantic es una libreria de python que permite la validacion de datos. FastAPI esta construida sobre ella

app = FastAPI()
#http://127.0.0.1:8000

# los decoradres modifican a la funcion a la que estan sujetos - en fastAPI no lo modifica sino lo que hace es registrar la funcion
@app.get("/")
def index():
    return 



root = tk.Tk()  # Create a root Tkinter instance for the application
root.title("Excel Translator")  # Set the title of the application window to "Excel Translator"

translator = Translator()  # Initialize the Google Translator for language translation
start_time = None  # Define a variable to store the starting time for the translation process

# Function to update the clock label with the elapsed time
def update_clock():
    elapsed_time = datetime.datetime.now() - start_time  # Calculate the elapsed time since the start of the translation process
    clock_label.config(text=f"Time Elapsed: {elapsed_time}")  # Update the clock label text to display the elapsed time
    root.after(1000, update_clock)  # Schedule the function to be called again after 1000 milliseconds (1 second)

# Function to handle the translation process
def translate_excel():
    global start_time  # Access the global variable for the start time of the translation process
    start_time = datetime.datetime.now()  # Store the current time as the start time for the translation process

    file_path = file_path_var.get()  # Get the selected file path from the corresponding variable

    if file_path:  # Check if a file path has been selected
        source_language = source_language_var.get()  # Retrieve the selected source language
        target_language = target_language_var.get()  # Retrieve the selected target language

        workbook = openpyxl.load_workbook(file_path)  # Load the selected Excel file using openpyxl

        for sheet_name in workbook.sheetnames:  # Iterate through the sheet names in the loaded workbook
            sheet = workbook[sheet_name]  # Access the current sheet
            for row in sheet.iter_rows():  # Iterate through the rows in the current sheet
                for cell in row:  # Iterate through the cells in the current row
                    if cell.value:  # Check if the cell has a value
                        translated_text = translator.translate(cell.value, src=source_language, dest=target_language)  # Translate the text using the selected source and target languages
                        cell.value = translated_text.text  # Update the cell value with the translated text

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx")  # Open a file dialog for saving the translated file
        if save_path:  # If a save path is selected
            workbook.save(save_path)  # Save the translated Excel file to the specified location

    update_clock()  # Start the clock to display the elapsed time

# Function to browse and select an Excel file
def browse_file():
    filename = filedialog.askopenfilename(initialdir="/", title="Select file", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))  # Open a file dialog to select an Excel file
    file_path_var.set(filename)  # Set the selected file path in the corresponding variable

file_path_var = tk.StringVar()  # Create a variable to store the selected file path
source_language_var = tk.StringVar(value='de')  # Create a variable to store the selected source language (default: German)
target_language_var = tk.StringVar(value='en')  # Create a variable to store the selected target language (default: English)

main_frame = ttk.Frame(root, padding=20)  # Create a main frame for organizing the UI elements with padding
main_frame.pack(expand=True, fill='both')  # Pack the main frame to expand and fill the entire window

# Create and pack various UI elements for the Excel Translator tool, such as labels, comboboxes, entry widgets, and buttons
title_label = ttk.Label(main_frame, text="Excel Translator Tool", font=("Helvetica", 16, "bold"))
title_label.pack(pady=10)

source_label = ttk.Label(main_frame, text="Source Language:")
source_label.pack(pady=5)

source_combobox = ttk.Combobox(main_frame, textvariable=source_language_var, values=['en', 'de'])
source_combobox.pack(pady=5)

target_label = ttk.Label(main_frame, text="Target Language:")
target_label.pack(pady=5)

target_combobox = ttk.Combobox(main_frame, textvariable=target_language_var, values=['en', 'de'])
target_combobox.pack(pady=5)

file_label = ttk.Label(main_frame, text="Selected Excel file:")
file_label.pack(pady=10)

file_entry = ttk.Entry(main_frame, textvariable=file_path_var, width=50)
file_entry.pack(pady=10)

browse_button = ttk.Button(main_frame, text="Browse", command=browse_file)
browse_button.pack(pady=10)

translate_button = ttk.Button(main_frame, text="Translate Excel File", command=translate_excel)
translate_button.pack(pady=10)

clock_label = ttk.Label(main_frame, text="Time Elapsed: 00:00:00")
clock_label.pack(pady=10)  # Pack the clock label with padding

root.mainloop()  # Start the Tkinter event loop to display the GUI and handle user interactions

