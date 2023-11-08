import tkinter as tk
from tkinter import ttk, filedialog
from googletrans import Translator
import openpyxl
import time

root = tk.Tk()
root.title("Excel Translator")

translator = Translator()
start_time = None

def translate_excel():
    global start_time
    start_time = time.time()

    file_path = file_path_var.get()

    if file_path:
        source_language = source_language_var.get()
        target_language = target_language_var.get()

        workbook = openpyxl.load_workbook(file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        translated_text = translator.translate(cell.value, src=source_language, dest=target_language)
                        cell.value = translated_text.text

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if save_path:
            workbook.save(save_path)

def browse_file():
    filename = filedialog.askopenfilename(initialdir="/", title="Select file", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    file_path_var.set(filename)

file_path_var = tk.StringVar()
source_language_var = tk.StringVar(value='de')
target_language_var = tk.StringVar(value='en')

main_frame = ttk.Frame(root, padding=20)
main_frame.pack(expand=True, fill='both')

title_label = ttk.Label(main_frame, text="Excel Translator Tool", font=("Helvetica", 16, "bold"))
title_label.pack(pady=10)

source_label = ttk.Label(main_frame, text="Source Language:")
source_label.pack(pady=5)

source_combobox = ttk.Combobox(main_frame, textvariable=source_language_var, values=['en', 'de'])
source_combobox.pack(pady=5)

target_label = ttk.Label(main_frame, text="Target Language:")
target_label.pack(pady=5)

target_combobox = ttk.Combobox(main_frame, textvariable=target_language_var, values=['en', 'de'])
target_combobox.pack(pady=5)

file_label = ttk.Label(main_frame, text="Selected Excel file:")
file_label.pack(pady=10)

file_entry = ttk.Entry(main_frame, textvariable=file_path_var, width=50)
file_entry.pack(pady=10)

browse_button = ttk.Button(main_frame, text="Browse", command=browse_file)
browse_button.pack(pady=10)

translate_button = ttk.Button(main_frame, text="Translate Excel File", command=translate_excel)
translate_button.pack(pady=10)

root.mainloop()